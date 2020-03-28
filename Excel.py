from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from init import *
from xlrd import open_workbook


def loadLists(textFile):
    with open(textFile) as file:
        list = file.readlines()
        file.close()
        list = [x[:-1] for x in list]
    return list


def absentCell(sheet, cell):
    # Add Red Color Cell Format
    redFill = PatternFill(start_color='F4CCCC',
                          end_color='F4CCCC',
                          fill_type='solid')
    sheet[cell].fill = redFill


def presentCell(sheet, cell):
    # Add Green Color Cell Format
    greenFill = PatternFill(start_color='D9EAD3',
                            end_color='D9EAD3',
                            fill_type='solid')
    sheet[cell].fill = greenFill


def lateCell(sheet, cell):
    # Add Yellow Color Cell Format
    yellowFill = PatternFill(start_color='FFF2CC',
                             end_color='FFF2CC',
                             fill_type='solid')
    sheet[cell].fill = yellowFill


def resetCell(sheet, cell):
    # Add White Color Cell Format
    whiteFill = PatternFill(start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid')
    sheet[cell].fill = whiteFill


def addKeyExcel(sheet):
    # Reset Top Cells
    for n in range(1, 5):
        cellLocation = 'A' + str(n)
        resetCell(sheet, cellLocation)

    # Add Key Colors and Labels
    presentCell(sheet, 'A2')
    absentCell(sheet, 'A3')
    lateCell(sheet, 'A4')
    sheet['A1'] = 'KEY'
    sheet['A1'].font = Font(bold=True)
    sheet['A2'] = 'Present'
    sheet['A2'].font = Font(bold=True)
    sheet['A3'] = 'Absent'
    sheet['A3'].font = Font(bold=True)
    sheet['A4'] = 'Late'
    sheet['A4'].font = Font(bold=True)


def addStudentNamesExcel(sheet):
    # Format and write Student Name subtitle
    sheet['A8'] = 'Student Names'
    sheet['A8'].font = Font(bold=True)
    # Write student names from init list
    for n in range(0, len(fullStudentNames)):
        cellLocation = 'A' + str(9 + n)
        sheet[cellLocation] = fullStudentNames[n]


def getRowNum(personToFind):
    startCellNum = 9
    for x in range(0, len(fullStudentNames)):
        # Find how many to go down from row 9 by comparing names + arrays
        if fullStudentNames[x].strip() == personToFind.strip():
            # Update row to go to
            startCellNum += x
    return startCellNum


def getColumnLetter(sheet):
    # Start column is B
    cellStartNum = ord('B')
    # Get date because column will correspond
    date = datetime.today().strftime('X%m/X%d')
    date = date.replace('X0', 'X').replace('X', '')
    columnFound = False
    # Compare current date to column date
    while not columnFound:
        currentCell = str(chr(cellStartNum)) + '8'
        # If found, return cell column Letter
        if sheet[currentCell].value == date:
            return cellStartNum
        else:
            cellStartNum += 1


def addDateExcel(sheet):
    # Get and format date
    date = datetime.today().strftime('X%m/X%d')
    date = date.replace('X0', 'X').replace('X', '')
    # character number for "B"
    cellStartNum = ord('B')
    # Flag boolean to exit loop
    emptyDateCell = False

    while not emptyDateCell:
        # Get Current cell location
        currentCell = str(chr(cellStartNum)) + '8'
        # If the date is already there, then you do not need to add another column
        if sheet[currentCell].value == date:
            break
        else:
            # # If cell is not empty, move over one cell horizontally
            if sheet[currentCell].value != None:
                cellStartNum += 1
            else:
                # If cell is empty, write the date
                sheet[currentCell] = date
                sheet[currentCell].font = Font(bold=True)
                emptyDateCell = True


def formatPageExcel(sheet):
    # Adds key, student names, and current date
    if sheet['A1'] != 'KEY':
        addKeyExcel(sheet)
    addStudentNamesExcel(sheet)
    addDateExcel(sheet)


def updatePresentPersonExcel(personToFind):
    # Change numerical values to cell value
    cellToPresent = chr(getColumnLetter(ws)) + str(getRowNum(personToFind))
    # Mark present
    presentCell(ws, cellToPresent)


def updateAbsentPersonExcel(personToFind):
    # Change numerical values to cell value
    cellToAbsent = chr(getColumnLetter(ws)) + str(getRowNum(personToFind))
    # Mark Absent
    absentCell(ws, cellToAbsent)


def updateLatePersonExcel(personToFind):
    # Change numerical values to cell value
    cellToAbsent = chr(getColumnLetter(ws)) + str(getRowNum(personToFind))
    # Mark Late
    lateCell(ws, cellToAbsent)


def markAbsentUnmarkedExcel():
    rowStart = 9
    for x in range(0, len(fullStudentNames)):
        cellToCheck = str(chr(getColumnLetter(ws))) + str(rowStart)
        if str(ws[cellToCheck].fill.start_color.index) not in '00D9EAD3':
            absentCell(ws, cellToCheck)
            rowStart += 1
        else:
            rowStart += 1
    wb.save("AttendanceExcel.xls")


try:
    fullStudentNames = loadLists("List Information/Full Student Names")
    wb = Workbook()
    ws = wb.active
    formatPageExcel(ws)
except Exception as e:
    print(e)
